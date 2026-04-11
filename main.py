"""
Point3 CRM - Flask 백엔드 서버
"""

import os
import json
import re
import uuid
import time
import csv
import io
import hashlib
import threading

from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from urllib.parse import quote
from functools import wraps

from flask import Flask, request, jsonify, render_template, Response, send_file, redirect, session
from flask_cors import CORS
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Flask 앱 초기화 ──
app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True
app.secret_key = os.environ.get("SECRET_KEY", "point3-crm-2026-secret")
CORS(app)

# ── 최고 관리자 설정 ──
SUPERADMIN_USERNAME = "leesk0130"
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "734818849350-qlm4r3mlbksfrv41hm38l78vscu29biu.apps.googleusercontent.com")

# ── Firestore 설정 ──
FIREBASE_PROJECT_ID = os.environ.get("FIREBASE_PROJECT_ID", "point3-salesmap99")
FIRESTORE_BASE = f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT_ID}/databases/(default)/documents"
FIREBASE_API_KEY = os.environ.get("FIREBASE_API_KEY", "AIzaSyA7u_44ljLdf5yxyihKO0qU51DkMZyiV_w")

# 로컬 폴백용 경로
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)


# ══════════════════════════════════════════
#  Firestore 헬퍼 함수
# ══════════════════════════════════════════

def _fs_to_dict(fields):
    """Firestore 필드 → 파이썬 dict 변환"""
    result = {}
    for k, v in fields.items():
        if not v or not isinstance(v, dict):
            result[k] = None
        elif "stringValue" in v:
            result[k] = v["stringValue"]
        elif "booleanValue" in v:
            result[k] = v["booleanValue"]
        elif "integerValue" in v:
            result[k] = int(v["integerValue"])
        elif "doubleValue" in v:
            result[k] = v["doubleValue"]
        elif "nullValue" in v:
            result[k] = None
        elif "arrayValue" in v:
            values = v["arrayValue"].get("values", [])
            result[k] = [_fs_to_dict(item.get("mapValue", {}).get("fields", {})) if "mapValue" in item else
                         item.get("stringValue", item.get("booleanValue", item.get("integerValue", "")))
                         for item in values]
        elif "mapValue" in v:
            result[k] = _fs_to_dict(v["mapValue"].get("fields", {}))
        else:
            result[k] = None
    return result


def _dict_to_fs(d):
    """파이썬 dict → Firestore 필드 변환"""
    fields = {}
    for k, v in d.items():
        if isinstance(v, bool):
            fields[k] = {"booleanValue": v}
        elif isinstance(v, int):
            fields[k] = {"integerValue": str(v)}
        elif isinstance(v, float):
            fields[k] = {"doubleValue": v}
        elif isinstance(v, str):
            fields[k] = {"stringValue": v}
        elif isinstance(v, list):
            values = []
            for item in v:
                if isinstance(item, dict):
                    values.append({"mapValue": {"fields": _dict_to_fs(item)}})
                elif isinstance(item, str):
                    values.append({"stringValue": item})
                else:
                    values.append({"stringValue": str(item)})
            fields[k] = {"arrayValue": {"values": values} if values else {"values": []}}
        elif v is None:
            fields[k] = {"nullValue": None}
        elif isinstance(v, dict):
            fields[k] = {"mapValue": {"fields": _dict_to_fs(v)}}
        else:
            fields[k] = {"stringValue": str(v)}
    return fields


def fs_get_collection(collection):
    """Firestore 컬렉션의 모든 문서 가져오기"""
    try:
        url = f"{FIRESTORE_BASE}/{collection}?key={FIREBASE_API_KEY}"
        resp = requests.get(url, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            docs = data.get("documents", [])
            return [_fs_to_dict(doc.get("fields", {})) for doc in docs]
        return []
    except Exception as e:
        print(f"[Firestore GET 오류] {collection}: {e}")
        return []


def fs_set_doc(collection, doc_id, data_dict):
    """Firestore 문서 생성/덮어쓰기"""
    try:
        url = f"{FIRESTORE_BASE}/{collection}/{doc_id}?key={FIREBASE_API_KEY}"
        body = {"fields": _dict_to_fs(data_dict)}
        resp = requests.patch(url, json=body, timeout=10)
        return resp.status_code == 200
    except Exception as e:
        print(f"[Firestore SET 오류] {collection}/{doc_id}: {e}")
        return False


def fs_delete_doc(collection, doc_id):
    """Firestore 문서 삭제"""
    try:
        url = f"{FIRESTORE_BASE}/{collection}/{doc_id}?key={FIREBASE_API_KEY}"
        resp = requests.delete(url, timeout=10)
        return resp.status_code in (200, 204)
    except Exception as e:
        print(f"[Firestore DEL 오류] {collection}/{doc_id}: {e}")
        return False


# ══════════════════════════════════════════
#  데이터 함수 (Firestore 기반)
# ══════════════════════════════════════════

_stores_cache = {"data": None, "ts": 0}

def load_stores():
    now = time.time()
    if _stores_cache["data"] is not None and (now - _stores_cache["ts"]) < 30:
        return _stores_cache["data"]
    result = fs_get_collection("stores")
    _stores_cache["data"] = result
    _stores_cache["ts"] = now
    return result

def _invalidate_cache():
    _stores_cache["data"] = None
    _stores_cache["ts"] = 0


# ══════════════════════════════════════════
#  활동 로그 시스템
# ══════════════════════════════════════════

def log_activity(team_name, username, action, store_name="", store_id="", detail=""):
    """활동 로그를 Firestore activity_logs 컬렉션에 기록"""
    try:
        log_entry = {
            "id": str(uuid.uuid4()),
            "teamName": team_name,
            "username": username,
            "action": action,
            "store_name": store_name,
            "store_id": store_id,
            "detail": detail,
            "timestamp": datetime.now().isoformat(),
        }
        fs_set_doc("activity_logs", log_entry["id"], log_entry)
    except Exception as e:
        print(f"[활동 로그 기록 오류] {e}")


def save_store(store):
    """매장 저장. 성공 시 True, 실패 시 False 반환"""
    result = fs_set_doc("stores", store["id"], store)
    _invalidate_cache()
    return result

def save_stores(stores):
    for s in stores:
        fs_set_doc("stores", s["id"], s)
    _invalidate_cache()

def delete_store_doc(store_id):
    result = fs_delete_doc("stores", store_id)
    _invalidate_cache()
    return result

def load_users():
    return fs_get_collection("users")

def save_user(user):
    fs_set_doc("users", user["id"], user)

def save_users(users):
    for u in users:
        fs_set_doc("users", u["id"], u)


# ══════════════════════════════════════════
#  인증 시스템
# ══════════════════════════════════════════

def hash_pw(pw):
    """새 비밀번호 해싱 (werkzeug/bcrypt 방식)"""
    return generate_password_hash(pw)


def _is_legacy_sha256(stored):
    """저장된 해시가 레거시 SHA256인지 판별 (64자 hex)"""
    return len(stored) == 64 and all(c in '0123456789abcdef' for c in stored)


def _verify_password(stored, plain):
    """비밀번호 검증: 레거시 SHA256 또는 werkzeug 해시 모두 지원"""
    if _is_legacy_sha256(stored):
        return hashlib.sha256(plain.encode()).hexdigest() == stored
    return check_password_hash(stored, plain)


def get_current_user():
    """세션에서 현재 로그인 유저 반환. 없으면 None"""
    uid = session.get("user_id")
    if not uid:
        return None
    users = load_users()
    for u in users:
        if u["id"] == uid:
            return u
    return None


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            if request.path.startswith("/api/"):
                return jsonify({"error": "로그인이 필요합니다."}), 401
            return redirect("/login")
        # 팀 미설정 → 팀 설정 페이지로
        if not user.get("teamName") and request.path not in ("/team-setup", "/auth/set-team", "/auth/logout"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "팀 설정이 필요합니다."}), 403
            return redirect("/team-setup")
        # 미승인 → 대기 페이지로
        if not user.get("isApproved") and request.path not in ("/pending", "/auth/me", "/auth/logout"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "승인 대기 중입니다."}), 403
            return redirect("/pending")
        return f(*args, **kwargs)
    return decorated


def superadmin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user or user.get("role") != "superadmin":
            return jsonify({"error": "권한이 없습니다."}), 403
        return f(*args, **kwargs)
    return decorated


# ── 초기 관리자 계정 자동 생성 ──
SUPERADMIN_DEFAULT_PW = os.environ.get("SUPERADMIN_PW", "Point3Admin!2026")

def ensure_superadmin():
    users = load_users()
    for u in users:
        if u["username"] == SUPERADMIN_USERNAME:
            return
    admin = {
        "id": str(uuid.uuid4()),
        "username": SUPERADMIN_USERNAME,
        "password": hash_pw(SUPERADMIN_DEFAULT_PW),
        "name": "관리자",
        "teamName": "point3",
        "isApproved": True,
        "role": "superadmin",
        "created_at": datetime.now().isoformat(),
    }
    users.append(admin)
    save_users(users)
    # 기존 매장에 teamName 할당
    stores = load_stores()
    changed = False
    for s in stores:
        if not s.get("teamName"):
            s["teamName"] = "Point3"
            changed = True
    if changed:
        save_stores(stores)



def extract_district(address):
    """주소에서 '구' 이름 추출 (예: '강남구'). 없으면 '기타' 반환"""
    match = re.search(r'\S+구', address)
    return match.group(0) if match else "기타"


def get_last_visit(store):
    """매장의 최근 방문일과 결과를 반환"""
    visits = store.get("visits", [])
    if visits:
        latest = max(visits, key=lambda v: v.get("date", ""))
        return latest.get("date", ""), latest.get("result", "")
    return "", ""


# ── 카카오 지오코딩 함수 ──
KAKAO_API_KEY = os.environ.get("KAKAO_API_KEY", "12a6d5580904db14be2b073e8e114a4f")

def geocode_address(address):
    """
    주소를 위도/경도로 변환 (카카오 로컬 API 사용)
    주소 검색 실패 시 키워드 검색으로 fallback
    """
    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}

    # 1차: 주소 검색
    try:
        url = "https://dapi.kakao.com/v2/local/search/address.json"
        resp = requests.get(url, headers=headers, params={"query": address}, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data.get("documents"):
            doc = data["documents"][0]
            return float(doc["y"]), float(doc["x"])
    except Exception as e:
        print(f"[지오코딩 오류] {address}: {e}")

    # 2차: 키워드 검색 (건물명 등)
    try:
        url = "https://dapi.kakao.com/v2/local/search/keyword.json"
        resp = requests.get(url, headers=headers, params={"query": address}, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data.get("documents"):
            doc = data["documents"][0]
            return float(doc["y"]), float(doc["x"])
    except Exception as e:
        print(f"[키워드 검색 오류] {address}: {e}")

    return None, None


# ══════════════════════════════════════════
#  인증 라우트
# ══════════════════════════════════════════

@app.route("/login")
def login_page():
    if get_current_user():
        return redirect("/")
    return render_template("login.html", google_client_id=GOOGLE_CLIENT_ID)


@app.route("/auth/signup", methods=["POST"])
def auth_signup():
    data = request.get_json()
    username = (data.get("username") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "")
    name = (data.get("name") or "").strip()

    if not username or not password or not name or not email:
        return jsonify({"error": "모든 필드를 입력하세요."}), 400
    if len(password) < 4:
        return jsonify({"error": "비밀번호는 4자 이상이어야 합니다."}), 400

    users = load_users()
    if any(u["username"] == username for u in users):
        return jsonify({"error": "이미 사용 중인 아이디입니다."}), 409

    user = {
        "id": str(uuid.uuid4()),
        "username": username,
        "email": email,
        "password": hash_pw(password),
        "name": name,
        "teamName": "",
        "isApproved": False,
        "role": "user",
        "created_at": datetime.now().isoformat(),
    }
    users.append(user)
    save_users(users)
    session["user_id"] = user["id"]
    return jsonify({"message": "회원가입 완료"})


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "")

    users = load_users()
    for u in users:
        if u["username"] == username and _verify_password(u.get("password", ""), password):
            # SHA256 → bcrypt 자동 마이그레이션
            if _is_legacy_sha256(u.get("password", "")):
                u["password"] = hash_pw(password)
                save_user(u)
            session["user_id"] = u["id"]
            return jsonify({
                "message": "로그인 성공",
                "needTeam": not u.get("teamName"),
                "needApproval": not u.get("isApproved"),
            })

    return jsonify({"error": "아이디 또는 비밀번호가 올바르지 않습니다."}), 401


@app.route("/auth/google", methods=["POST"])
def auth_google():
    data = request.get_json()
    credential = data.get("credential")
    if not credential:
        return jsonify({"error": "Google 인증 정보가 없습니다."}), 400

    try:
        from google.oauth2 import id_token
        from google.auth.transport import requests as google_requests
        idinfo = id_token.verify_oauth2_token(credential, google_requests.Request(), GOOGLE_CLIENT_ID)
        email = idinfo["email"].lower()
        name = idinfo.get("name", email.split("@")[0])
    except Exception as e:
        print(f"[Google Auth Error] {str(e)}")
        return jsonify({"error": "Google 인증에 실패했습니다. 다시 시도해주세요."}), 401

    users = load_users()
    # 기존 유저 찾기
    for u in users:
        if u.get("email") == email or u.get("username") == email:
            session["user_id"] = u["id"]
            return jsonify({"message": "Google 로그인 성공", "needTeam": not u.get("teamName")})

    # 새 유저 생성
    user = {
        "id": str(uuid.uuid4()),
        "username": email,
        "password": "",
        "name": name,
        "email": email,
        "teamName": "",
        "isApproved": False,
        "role": "user",
        "created_at": datetime.now().isoformat(),
    }
    users.append(user)
    save_users(users)
    session["user_id"] = user["id"]
    return jsonify({"message": "Google 회원가입 완료", "needTeam": True})


@app.route("/auth/logout")
def auth_logout():
    session.clear()
    return redirect("/login")


@app.route("/auth/me")
def auth_me():
    user = get_current_user()
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    return jsonify({
        "id": user["id"], "name": user["name"], "username": user["username"],
        "teamName": user.get("teamName", ""), "isApproved": user.get("isApproved", False),
        "role": user.get("role", "user"),
    })


@app.route("/team-setup")
def team_setup_page():
    user = get_current_user()
    if not user:
        return redirect("/login")
    if user.get("teamName"):
        return redirect("/pending" if not user.get("isApproved") else "/")
    return render_template("team-setup.html", team_name="")


@app.route("/auth/set-team", methods=["POST"])
def auth_set_team():
    user = get_current_user()
    if not user:
        return jsonify({"error": "로그인 필요"}), 401
    data = request.get_json()
    team = (data.get("teamName") or "").strip()
    if not team:
        return jsonify({"error": "팀 이름을 입력하세요."}), 400

    users = load_users()
    for u in users:
        if u["id"] == user["id"]:
            u["teamName"] = team
            break
    save_users(users)
    return jsonify({"message": "팀 설정 완료"})


@app.route("/pending")
def pending_page():
    user = get_current_user()
    if not user:
        return redirect("/login")
    if user.get("isApproved"):
        return redirect("/")
    return render_template("pending.html", team_name=user.get("teamName", "") if user else "")


@app.route("/approve")
@login_required
def approve_page():
    user = get_current_user()
    if user.get("role") != "superadmin":
        return redirect("/")
    return render_template("approve.html", team_name=user.get("teamName", ""))


@app.route("/teams")
@login_required
def teams_page():
    user = get_current_user()
    if user.get("role") != "superadmin":
        return redirect("/")
    return render_template("teams.html", team_name=user.get("teamName", ""))


# ── 관리자 API ──

@app.route("/api/admin/pending-users")
@superadmin_required
def admin_pending_users():
    users = load_users()
    pending = [{"id": u["id"], "name": u["name"], "username": u["username"],
                "email": u.get("email", ""), "teamName": u.get("teamName", ""), "created_at": u.get("created_at", "")}
               for u in users if not u.get("isApproved") and u.get("teamName")]
    return jsonify(pending)


@app.route("/api/admin/approved-users")
@superadmin_required
def admin_approved_users():
    users = load_users()
    approved = [{"id": u["id"], "name": u["name"], "username": u["username"],
                 "email": u.get("email", ""), "teamName": u.get("teamName", ""), "role": u.get("role", "user")}
                for u in users if u.get("isApproved")]
    return jsonify(approved)


@app.route("/api/admin/approve", methods=["POST"])
@superadmin_required
def admin_approve():
    data = request.get_json()
    uid = data.get("userId")
    users = load_users()
    for u in users:
        if u["id"] == uid:
            u["isApproved"] = True
            save_users(users)
            return jsonify({"message": f"{u['name']} 승인 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/admin/reject", methods=["POST"])
@superadmin_required
def admin_reject():
    data = request.get_json()
    uid = data.get("userId")
    if uid and fs_delete_doc("users", uid):
        return jsonify({"message": "거절 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/admin/delete-user", methods=["POST"])
@superadmin_required
def admin_delete_user():
    """승인된 사용자 삭제"""
    data = request.get_json()
    uid = data.get("userId")
    if uid and fs_delete_doc("users", uid):
        return jsonify({"message": "삭제 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/admin/deactivate-user", methods=["POST"])
@superadmin_required
def admin_deactivate_user():
    """사용자 비활성화 (삭제가 아닌 isApproved=False 처리)"""
    data = request.get_json()
    uid = data.get("userId")
    if not uid:
        return jsonify({"error": "userId가 필요합니다."}), 400
    users = load_users()
    for u in users:
        if u["id"] == uid:
            if u.get("role") == "superadmin":
                return jsonify({"error": "관리자는 비활성화할 수 없습니다."}), 400
            u["isApproved"] = False
            save_user(u)
            return jsonify({"message": f"{u['name']} 비활성화 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/admin/reactivate-user", methods=["POST"])
@superadmin_required
def admin_reactivate_user():
    """비활성화된 사용자 재활성화"""
    data = request.get_json()
    uid = data.get("userId")
    if not uid:
        return jsonify({"error": "userId가 필요합니다."}), 400
    users = load_users()
    for u in users:
        if u["id"] == uid:
            u["isApproved"] = True
            save_user(u)
            return jsonify({"message": f"{u['name']} 재활성화 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/users/<username>/role", methods=["PUT"])
@superadmin_required
def update_user_role(username):
    """팀원 역할 변경 (관리자/팀원/뷰어)"""
    data = request.get_json()
    new_role = data.get("role", "").strip()
    valid_roles = ["admin", "user", "viewer"]
    if new_role not in valid_roles:
        return jsonify({"error": "유효하지 않은 역할입니다."}), 400
    users = load_users()
    for u in users:
        if u["username"] == username:
            if u.get("role") == "superadmin":
                return jsonify({"error": "최고 관리자의 역할은 변경할 수 없습니다."}), 400
            u["role"] = new_role
            save_user(u)
            return jsonify({"message": f"역할 변경 완료: {new_role}", "role": new_role})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


@app.route("/api/team/invite-link")
@superadmin_required
def team_invite_link():
    """팀 초대 링크 생성"""
    user = get_current_user()
    team = user.get("teamName", "")
    token = hashlib.sha256(f"{team}:{app.secret_key}:invite".encode()).hexdigest()[:16]
    base_url = request.host_url.rstrip("/")
    link = f"{base_url}/login?invite={token}&team={quote(team)}"
    return jsonify({"link": link, "team": team})


@app.route("/api/team/members-stats")
@superadmin_required
def team_members_stats():
    """팀원별 활동 통계 (최근 7일 활동, 담당 가맹점 수)"""
    from datetime import timedelta
    user = get_current_user()
    team = user.get("teamName", "")
    users_list = load_users()
    team_members = [u for u in users_list if u.get("teamName") == team]
    stores = load_stores()
    my_stores = [s for s in stores if s.get("teamName") == team]
    today = datetime.now().date()
    seven_days_ago = (today - timedelta(days=7)).strftime("%Y-%m-%d")
    result = []
    for member in team_members:
        mname = member.get("name", "")
        recent_7d = 0
        total_visits = 0
        assigned_stores = 0
        for s in my_stores:
            visits = s.get("visits") or []
            member_visited = False
            for v in visits:
                v_author = v.get("author", "") or v.get("created_by", "")
                is_mine = (v_author == mname) or (len(team_members) == 1)
                if is_mine:
                    total_visits += 1
                    member_visited = True
                    vdate = v.get("date", "")
                    if vdate >= seven_days_ago:
                        recent_7d += 1
            if member_visited:
                assigned_stores += 1
        result.append({
            "id": member["id"], "username": member.get("username", ""),
            "name": mname, "email": member.get("email", ""),
            "role": member.get("role", "user"),
            "isApproved": member.get("isApproved", False),
            "teamName": member.get("teamName", ""),
            "created_at": member.get("created_at", ""),
            "recent_7d": recent_7d, "total_visits": total_visits,
            "assigned_stores": assigned_stores,
        })
    return jsonify(result)


@app.route("/api/stores/check-duplicate")
@login_required
def check_duplicate_store():
    """가맹점명 중복 체크"""
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"exists": False})
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    for s in stores:
        if s.get("teamName") == team and s.get("name", "").strip() == name:
            return jsonify({"exists": True, "store": {"name": s["name"], "address": s.get("address", "")}})
    return jsonify({"exists": False})


@app.route("/api/upload-excel-preview", methods=["POST"])
@login_required
def upload_excel_preview():
    """엑셀 파일 미리보기 - 실제 등록하지 않고 파싱 결과만 반환"""
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": ".xlsx 파일만 업로드 가능합니다."}), 400
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"엑셀 파일 읽기 실패: {str(e)}"}), 400
    headers = [cell.value for cell in ws[1]]
    required = ["가맹점명", "주소"]
    for col in required:
        if col not in headers:
            return jsonify({"error": f"필수 컬럼 '{col}'이(가) 없습니다."}), 400
    col_map = {}
    for idx, h in enumerate(headers):
        col_map[h] = idx
    user = get_current_user()
    team = user.get("teamName", "")
    existing_stores = load_stores()
    existing_names = {s.get("name", "").strip() for s in existing_stores if s.get("teamName") == team}
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name_val = row[col_map["가맹점명"]] if col_map.get("가맹점명") is not None else None
        addr_val = row[col_map["주소"]] if col_map.get("주소") is not None else None
        if not name_val or not addr_val:
            continue
        memo = ""
        if "메모" in col_map and col_map["메모"] < len(row):
            memo = row[col_map["메모"]] or ""
        is_dup = str(name_val).strip() in existing_names
        rows.append({"row": row_idx, "name": str(name_val), "address": str(addr_val),
                     "memo": str(memo), "duplicate": is_dup})
    return jsonify({"rows": rows, "total": len(rows)})


@app.route("/api/stores/recent")
@login_required
def recent_stores():
    """최근 등록한 가맹점 5개"""
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my_stores = [s for s in stores if s.get("teamName") == team]
    my_stores.sort(key=lambda s: s.get("created_at", ""), reverse=True)
    return jsonify(my_stores[:5])


@app.route("/api/admin/all-teams")
@superadmin_required
def admin_all_teams():
    """전체 팀 목록 + 통계 (슈퍼관리자용)"""
    users = load_users()
    stores = load_stores()

    teams = {}
    for u in users:
        tn = u.get("teamName", "").strip()
        if not tn:
            continue
        if tn not in teams:
            teams[tn] = {"teamName": tn, "members": 0, "approved": 0, "pending": 0,
                         "inactive": 0, "stores": 0, "visits": 0, "users": []}
        t = teams[tn]
        t["members"] += 1
        if u.get("isApproved"):
            t["approved"] += 1
        elif u.get("teamName"):
            t["pending"] += 1
        t["users"].append({
            "id": u.get("id", ""), "name": u.get("name", ""), "username": u.get("username", ""),
            "email": u.get("email", ""), "role": u.get("role", "user"),
            "isApproved": u.get("isApproved", False),
            "created_at": u.get("created_at", "")
        })

    for s in stores:
        tn = s.get("teamName", "").strip()
        if tn in teams:
            teams[tn]["stores"] += 1
            visits = s.get("visits", [])
            if isinstance(visits, list):
                teams[tn]["visits"] += len(visits)

    result = sorted(teams.values(), key=lambda x: x["members"], reverse=True)
    return jsonify(result)


@app.route("/api/admin/delete-team", methods=["POST"])
@superadmin_required
def admin_delete_team():
    """팀 전체 삭제 (멤버 + 매장 모두 삭제)"""
    data = request.get_json()
    team_name = (data.get("teamName") or "").strip()
    if not team_name:
        return jsonify({"error": "teamName이 필요합니다."}), 400
    if team_name.lower() == "point3":
        return jsonify({"error": "기본 팀은 삭제할 수 없습니다."}), 400

    users = load_users()
    stores = load_stores()
    deleted_users = 0
    deleted_stores = 0

    for u in users:
        if u.get("teamName", "").strip() == team_name and u.get("role") != "superadmin":
            fs_delete_doc("users", u["id"])
            deleted_users += 1

    for s in stores:
        if s.get("teamName", "").strip() == team_name:
            fs_delete_doc("stores", s["id"])
            deleted_stores += 1

    _invalidate_cache()
    return jsonify({"message": f"팀 '{team_name}' 삭제 완료 (멤버 {deleted_users}명, 매장 {deleted_stores}개)"})


@app.route("/api/admin/update-team", methods=["POST"])
@superadmin_required
def admin_update_team():
    """사용자 팀명 수정"""
    data = request.get_json()
    uid = data.get("userId")
    new_team = (data.get("teamName") or "").strip()
    if not uid or not new_team:
        return jsonify({"error": "userId와 teamName이 필요합니다."}), 400
    users = load_users()
    for u in users:
        if u["id"] == uid:
            u["teamName"] = new_team
            save_user(u)
            return jsonify({"message": f"팀명 '{new_team}'으로 변경 완료"})
    return jsonify({"error": "유저를 찾을 수 없습니다."}), 404


# ══════════════════════════════════════════
#  페이지 라우트 (로그인+승인 필수)
# ══════════════════════════════════════════

@app.route("/")
@login_required
def index():
    user = get_current_user()
    return render_template("map.html", team_name=user.get("teamName", ""))


@app.route("/admin")
@login_required
def admin_page():
    user = get_current_user()
    return render_template("admin.html", is_superadmin=(user.get("role") == "superadmin"), team_name=user.get("teamName", ""))


@app.route("/dashboard")
@login_required
def dashboard():
    user = get_current_user()
    return render_template("dashboard.html", team_name=user.get("teamName", ""))


@app.route("/calendar")
@login_required
def calendar():
    user = get_current_user()
    return render_template("calendar.html", team_name=user.get("teamName", ""))


@app.route("/stores")
@login_required
def stores_page():
    user = get_current_user()
    return render_template("stores.html", team_name=user.get("teamName", ""))


# ══════════════════════════════════════════
#  API 라우트 - 매장 CRUD
# ══════════════════════════════════════════

@app.route("/api/search", methods=["GET"])
@login_required
def search_stores():
    """가맹점 통합 검색 (이름+주소+지역구). 초성 검색 지원"""
    user = get_current_user()
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])

    stores = load_stores()
    team = user.get("teamName", "")
    my_stores = [s for s in stores if s.get("teamName") == team]

    q_lower = q.lower()
    results = []
    for s in my_stores:
        name = (s.get("name") or "").lower()
        address = (s.get("address") or "").lower()
        district = (s.get("district") or "").lower()
        if q_lower in name or q_lower in address or q_lower in district:
            visits = s.get("visits", [])
            status = "미컨택"
            if visits:
                latest = max(visits, key=lambda v: v.get("date", ""))
                raw = latest.get("result", "미컨택")
                status_map = {'미방문':'미컨택','부재중':'미컨택','계약성사':'미팅완료','실패':'미컨택','기타':'미컨택'}
                valid = ['미컨택','명함전달','미팅대기','미팅완료']
                status = status_map.get(raw, raw if raw in valid else '미컨택')
            results.append({
                "id": s.get("id"),
                "name": s.get("name"),
                "address": s.get("address"),
                "district": s.get("district"),
                "status": status,
                "lat": s.get("lat"),
                "lng": s.get("lng"),
            })

    return jsonify(results[:20])


@app.route("/api/stores", methods=["GET"])
@login_required
def get_stores():
    """현재 유저의 팀 매장 목록 조회"""
    user = get_current_user()
    stores = load_stores()
    team = user.get("teamName", "")
    my_stores = [s for s in stores if s.get("teamName") == team]
    return jsonify(my_stores)


@app.route("/api/stores", methods=["POST"])
@login_required
def add_store():
    """새 매장 추가"""
    user = get_current_user()
    data = request.get_json()

    if not data or not data.get("name") or not data.get("address"):
        return jsonify({"error": "매장명과 주소는 필수입니다."}), 400

    address = data.get("address", "")
    store = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", ""),
        "address": address,
        "lat": data.get("lat"),
        "lng": data.get("lng"),
        "district": extract_district(address),
        "memo": data.get("memo", ""),
        "notes": data.get("notes", ""),
        "contact_email": data.get("contact_email", False),
        "contact_linkedin": data.get("contact_linkedin", False),
        "contact_remember": data.get("contact_remember", False),
        "contact_intro": data.get("contact_intro", False),
        "showOnMap": data.get("showOnMap", True),
        "website": data.get("website", ""),
        "visits": [],
        "teamName": user.get("teamName", ""),
        "created_at": datetime.now().isoformat(),
    }

    # 위도/경도가 없으면 자동 지오코딩
    if store["lat"] is None or store["lng"] is None:
        lat, lng = geocode_address(store["address"])
        if lat is None or lng is None:
            return jsonify({"error": "주소를 지도 좌표로 변환할 수 없습니다. 정확한 주소인지 확인해주세요."}), 400
        store["lat"] = lat
        store["lng"] = lng

    if not save_store(store):
        return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
    log_activity(user.get("teamName", ""), user.get("name", ""), "가맹점 등록",
                 store.get("name", ""), store["id"], f"주소: {store.get('address', '')}")
    return jsonify(store), 201


@app.route("/api/stores/<store_id>", methods=["PUT"])
@login_required
def update_store(store_id):
    """매장 정보 수정"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    data = request.get_json()
    stores = load_stores()

    # 해당 ID의 매장 찾기
    for i, store in enumerate(stores):
        if store["id"] == store_id:
            # 팀 소유권 검증
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            # 전달된 필드만 업데이트
            for key in ["name", "address", "lat", "lng", "memo", "notes",
                         "contact_email", "contact_linkedin", "contact_remember",
                         "contact_intro", "showOnMap", "website"]:
                if key in data:
                    stores[i][key] = data[key]
            # 주소가 변경되면 district 재추출
            if "address" in data:
                stores[i]["district"] = extract_district(data["address"])
            if not save_store(stores[i]):
                return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
            changed_keys = [k for k in data if k in ["name","address","memo","notes","contact_email","contact_linkedin","contact_remember","contact_intro","showOnMap","website"]]
            detail = ", ".join(changed_keys) + " 변경" if changed_keys else "정보 수정"
            log_activity(user_team, user.get("name", ""), "가맹점 수정",
                         stores[i].get("name", ""), store_id, detail)
            return jsonify(stores[i])

    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


@app.route("/api/stores/<store_id>", methods=["DELETE"])
@login_required
def delete_store(store_id):
    """매장 삭제"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    stores = load_stores()
    for store in stores:
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            break
    else:
        return jsonify({"error": "매장을 찾을 수 없습니다."}), 404

    if delete_store_doc(store_id):
        log_activity(user_team, user.get("name", ""), "가맹점 삭제",
                     store.get("name", ""), store_id, "")
        return jsonify({"message": "삭제 완료"}), 200
    return jsonify({"error": "삭제 중 오류가 발생했습니다."}), 500


# ══════════════════════════════════════════
#  API 라우트 - 엑셀 업로드
# ══════════════════════════════════════════

@app.route("/api/upload-excel", methods=["POST"])
@login_required
def upload_excel():
    """
    엑셀(.xlsx) 파일 업로드 후 매장 일괄 등록
    엑셀 컬럼: 가맹점명, 주소, 메모
    각 주소를 카카오 API로 지오코딩
    """
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": ".xlsx 파일만 업로드 가능합니다."}), 400

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"엑셀 파일 읽기 실패: {str(e)}"}), 400

    # 헤더 행 읽기 (첫 번째 행)
    headers = [cell.value for cell in ws[1]]
    required = ["가맹점명", "주소"]
    for col in required:
        if col not in headers:
            return jsonify({"error": f"필수 컬럼 '{col}'이(가) 없습니다."}), 400

    # 컬럼 인덱스 매핑
    col_map = {}
    for idx, h in enumerate(headers):
        col_map[h] = idx

    user = get_current_user()
    team = user.get("teamName", "") if user else ""
    added = []
    errors = []

    # 데이터 행 순회 (2번째 행부터)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[col_map["가맹점명"]] if col_map.get("가맹점명") is not None else None
        address = row[col_map["주소"]] if col_map.get("주소") is not None else None

        # 빈 행 건너뛰기
        if not name or not address:
            continue

        memo = ""
        if "메모" in col_map and col_map["메모"] < len(row):
            memo = row[col_map["메모"]] or ""

        # 카카오 지오코딩
        address_str = str(address)
        lat, lng = geocode_address(address_str)

        store = {
            "id": str(uuid.uuid4()),
            "name": str(name),
            "address": address_str,
            "lat": lat,
            "lng": lng,
            "district": extract_district(address_str),
            "memo": str(memo),
            "visits": [],
            "teamName": team,
            "created_at": datetime.now().isoformat(),
        }

        if lat is None or lng is None:
            errors.append(f"행 {row_idx}: '{address}' 지오코딩 실패")

        save_store(store)
        added.append(store)

    if added:
        log_activity(team, user.get("name", ""), "엑셀 업로드",
                     "", "", f"{len(added)}개 매장 일괄 등록")
    return jsonify({
        "message": f"{len(added)}개 매장 추가 완료",
        "added": added,
        "errors": errors,
    }), 201


# ══════════════════════════════════════════
#  API 라우트 - 방문 기록 (Visit Logs)
# ══════════════════════════════════════════

@app.route("/api/stores/<store_id>/visits", methods=["GET"])
@login_required
def get_visits(store_id):
    """매장의 방문 기록 조회"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    stores = load_stores()
    for store in stores:
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            return jsonify(store.get("visits", [])), 200
    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


@app.route("/api/stores/<store_id>/visits", methods=["POST"])
@login_required
def add_visit(store_id):
    """매장에 방문 기록 추가"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    data = request.get_json()
    stores = load_stores()

    for i, store in enumerate(stores):
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            visit = {
                "id": str(uuid.uuid4()),
                "date": data.get("date", datetime.now().strftime("%Y-%m-%d")),
                "result": data.get("result", ""),
                "memo": data.get("memo", ""),
                "created_at": datetime.now().isoformat(),
            }
            # visits 필드가 없는 기존 데이터 호환
            if "visits" not in stores[i]:
                stores[i]["visits"] = []
            stores[i]["visits"].append(visit)
            if not save_store(stores[i]):
                return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
            log_activity(user_team, user.get("name", ""), "상태 변경",
                         store.get("name", ""), store_id,
                         f"{visit.get('result', '')} ({visit.get('date', '')})")
            return jsonify(visit), 201

    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


@app.route("/api/stores/<store_id>/visits/<visit_id>/date", methods=["PATCH"])
@login_required
def update_visit_date(store_id, visit_id):
    """방문 기록의 날짜 변경 (캘린더 드래그 이동용)"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    data = request.get_json()
    new_date = data.get("date")
    if not new_date:
        return jsonify({"error": "date 필드가 필요합니다."}), 400

    _invalidate_cache()
    stores = load_stores()

    for i, store in enumerate(stores):
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            visits = store.get("visits", [])
            for v in visits:
                if v.get("id") == visit_id:
                    v["date"] = new_date
                    stores[i]["visits"] = visits
                    if not save_store(stores[i]):
                        return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
                    return jsonify({"message": "날짜 변경 완료", "visit": v}), 200
            return jsonify({"error": "방문 기록을 찾을 수 없습니다."}), 404

    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


@app.route("/api/stores/<store_id>/visits/<visit_id>", methods=["DELETE"])
@login_required
def delete_visit(store_id, visit_id):
    """매장의 방문 기록 삭제"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    # 캐시 무효화 후 최신 데이터로 작업
    _invalidate_cache()
    stores = load_stores()

    for i, store in enumerate(stores):
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            visits = store.get("visits", [])
            original_len = len(visits)
            # 깊은 복사로 캐시 오염 방지
            updated_store = dict(store)
            updated_store["visits"] = [v for v in visits if v.get("id") != visit_id]

            if len(updated_store["visits"]) == original_len:
                return jsonify({"error": "방문 기록을 찾을 수 없습니다."}), 404

            if not save_store(updated_store):
                return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
            log_activity(user_team, user.get("name", ""), "방문기록 삭제",
                         store.get("name", ""), store_id, "")
            return jsonify({"message": "방문 기록 삭제 완료"}), 200

    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


# ══════════════════════════════════════════
#  API 라우트 - 통계
# ══════════════════════════════════════════

@app.route("/api/stats", methods=["GET"])
@login_required
def get_stats():
    from datetime import timedelta
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]
    total = len(my)
    visited = sum(1 for s in my if s.get("visits"))
    status_counts = {}
    recent_7d = 0
    today = datetime.now().date()
    for s in my:
        vs = s.get("visits") or []
        if vs:
            latest = max(vs, key=lambda v: v.get("date", ""))
            r = latest.get("result", "기타") or "기타"
            status_counts[r] = status_counts.get(r, 0) + 1
            for v in vs:
                try:
                    vd = datetime.strptime(v.get("date", ""), "%Y-%m-%d").date()
                    if (today - vd).days <= 7:
                        recent_7d += 1
                except:
                    pass
    return jsonify({"total": total, "visited": visited, "not_visited": total - visited,
                    "status_counts": status_counts, "recent_7d_visits": recent_7d})


# ══════════════════════════════════════════
#  API 라우트 - 지역구 목록
# ══════════════════════════════════════════

@app.route("/api/districts", methods=["GET"])
@login_required
def get_districts():
    """모든 매장의 지역구 목록과 개수 조회"""
    stores = load_stores()
    district_counts = {}
    for store in stores:
        district = store.get("district", "기타")
        district_counts[district] = district_counts.get(district, 0) + 1

    result = [{"name": name, "count": count} for name, count in district_counts.items()]
    result.sort(key=lambda x: x["count"], reverse=True)
    return jsonify(result)


# ══════════════════════════════════════════
#  API 라우트 - 주소 지오코딩
# ══════════════════════════════════════════

@app.route("/api/geocode", methods=["GET"])
@login_required
def geocode():
    """주소를 위도/경도로 변환"""
    address = request.args.get("address", "")
    if not address:
        return jsonify({"error": "주소를 입력하세요."}), 400

    lat, lng = geocode_address(address)
    if lat is None:
        return jsonify({"error": "지오코딩 실패: 주소를 찾을 수 없습니다."}), 404

    return jsonify({"lat": lat, "lng": lng})


# ══════════════════════════════════════════
#  API 라우트 - 주소 자동완성 검색
# ══════════════════════════════════════════

@app.route("/api/address-search", methods=["GET"])
@login_required
def address_search():
    """주소 자동완성 (카카오 로컬 API 사용)"""
    query = request.args.get("q", "").strip()
    if not query or len(query) < 2:
        return jsonify([])

    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}

    try:
        url = "https://dapi.kakao.com/v2/local/search/keyword.json"
        resp = requests.get(url, headers=headers, params={"query": query, "size": 7}, timeout=5)
        resp.raise_for_status()
        data = resp.json()
        results = []
        for item in data.get("documents", []):
            results.append({
                "display_name": item.get("address_name", item.get("place_name", "")),
                "lat": float(item["y"]),
                "lng": float(item["x"]),
            })
        return jsonify(results)
    except Exception as e:
        print(f"[주소검색 오류] {query}: {e}")
        return jsonify([])


# ══════════════════════════════════════════
#  API 라우트 - 즐겨찾기 토글
# ══════════════════════════════════════════

@app.route("/api/stores/<store_id>/star", methods=["POST"])
@login_required
def toggle_star(store_id):
    """매장 즐겨찾기 토글 (starred 필드 true/false)"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    stores = load_stores()

    for i, store in enumerate(stores):
        if store["id"] == store_id:
            if store.get("teamName") != user_team:
                return jsonify({"error": "권한이 없습니다."}), 403
            current = stores[i].get("starred", False)
            stores[i]["starred"] = not current
            if not save_store(stores[i]):
                return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500
            return jsonify({
                "id": store_id,
                "starred": stores[i]["starred"],
            })

    return jsonify({"error": "매장을 찾을 수 없습니다."}), 404


# ══════════════════════════════════════════
#  API 라우트 - 일괄 삭제
# ══════════════════════════════════════════

@app.route("/api/stores/bulk-delete", methods=["POST"])
@login_required
def bulk_delete_stores():
    """매장 일괄 삭제 (body: {"ids": ["id1", "id2", ...]})"""
    user = get_current_user()
    user_team = user.get("teamName", "")
    data = request.get_json()
    if not data or not isinstance(data.get("ids"), list):
        return jsonify({"error": "ids 배열이 필요합니다."}), 400

    ids_to_delete = set(data["ids"])
    if not ids_to_delete:
        return jsonify({"error": "삭제할 ID가 없습니다."}), 400

    # 팀 소유 매장만 삭제 허용
    stores = load_stores()
    my_store_ids = {s["id"] for s in stores if s.get("teamName") == user_team}
    ids_to_delete = ids_to_delete & my_store_ids

    deleted_count = 0
    for sid in ids_to_delete:
        if delete_store_doc(sid):
            deleted_count += 1

    if deleted_count == 0:
        return jsonify({"error": "일치하는 매장이 없습니다."}), 404
    return jsonify({
        "message": f"{deleted_count}개 매장 삭제 완료",
        "deleted_count": deleted_count,
    }), 200


# ══════════════════════════════════════════
#  API 라우트 - CSV 내보내기
# ══════════════════════════════════════════

@app.route("/api/export/csv")
@login_required
def export_csv():
    """가맹점 목록을 BOM 포함 UTF-8 CSV로 내보내기. 필터 파라미터 지원"""
    stores = load_stores()
    user = get_current_user()
    team = user.get("teamName", "")
    stores = [s for s in stores if s.get("teamName") == team]

    # 필터: ids 파라미터로 특정 매장만 내보내기
    f_ids = request.args.get("ids", "")
    if f_ids:
        id_set = set(f_ids.split(","))
        stores = [s for s in stores if s.get("id") in id_set]

    output = io.StringIO()
    output.write('\ufeff')

    writer = csv.writer(output)
    writer.writerow([
        "가맹점명", "주소", "지역구", "메모", "비고사항",
        "이메일", "링크드인", "리멤버", "소개",
        "지도표시", "최근방문일", "최근방문결과", "총방문수",
        "즐겨찾기", "등록일",
    ])

    for store in stores:
        visits = store.get("visits", [])
        last_date, last_result = get_last_visit(store)

        writer.writerow([
            store.get("name", ""),
            store.get("address", ""),
            store.get("district", ""),
            store.get("memo", ""),
            store.get("notes", ""),
            "Y" if store.get("contact_email") else "N",
            "Y" if store.get("contact_linkedin") else "N",
            "Y" if store.get("contact_remember") else "N",
            "Y" if store.get("contact_intro") else "N",
            "Y" if store.get("showOnMap", True) else "N",
            last_date,
            last_result,
            len(visits),
            "Y" if store.get("starred", False) else "N",
            store.get("created_at", ""),
        ])

    filename = f"stores_{datetime.now().strftime('%Y%m%d')}.csv"
    encoded_filename = quote(f"가맹점목록_{datetime.now().strftime('%Y%m%d')}.csv")

    response = Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename=\"{filename}\"; "
        f"filename*=UTF-8''{encoded_filename}"
    )
    return response


# ══════════════════════════════════════════
#  API 라우트 - 엑셀 내보내기
# ══════════════════════════════════════════

@app.route("/api/export/excel")
@login_required
def export_excel():
    """가맹점 목록을 .xlsx로 내보내기 (시트1: 목록, 시트2: 방문기록). 필터 파라미터 지원"""
    stores = load_stores()
    user = get_current_user()
    team = user.get("teamName", "")
    stores = [s for s in stores if s.get("teamName") == team]

    # 필터 파라미터 적용
    f_status = request.args.get("status", "")
    f_district = request.args.get("district", "")
    f_search = request.args.get("search", "").lower()
    f_ids = request.args.get("ids", "")

    if f_ids:
        id_set = set(f_ids.split(","))
        stores = [s for s in stores if s.get("id") in id_set]
    else:
        if f_status:
            def _get_export_status(s):
                visits = s.get("visits", [])
                if not visits:
                    return "미컨택"
                latest = max(visits, key=lambda v: v.get("date", ""))
                r = latest.get("result", "미컨택")
                m = {"미방문": "미컨택", "부재중": "미컨택", "계약성사": "미팅완료", "실패": "미컨택", "기타": "미컨택"}
                valid = ["미컨택", "명함전달", "미팅대기", "미팅완료"]
                return m.get(r, r if r in valid else "미컨택")
            stores = [s for s in stores if _get_export_status(s) == f_status]
        if f_district:
            stores = [s for s in stores if (s.get("district") or "기타") == f_district]
        if f_search:
            stores = [s for s in stores if f_search in f"{s.get('name','')} {s.get('address','')} {s.get('memo','')}".lower()]

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # 시트1: 가맹점 목록
    ws1 = wb.active
    ws1.title = "가맹점 목록"
    headers1 = ["가맹점명", "주소", "지역구", "메모", "비고사항",
                 "이메일", "링크드인", "리멤버", "소개", "지도표시",
                 "최근방문일", "최근방문결과", "총방문수", "즐겨찾기", "등록일"]
    ws1.append(headers1)
    style_header_row(ws1, len(headers1))

    for store in stores:
        visits = store.get("visits", [])
        last_date, last_result = get_last_visit(store)

        ws1.append([
            store.get("name", ""),
            store.get("address", ""),
            store.get("district", ""),
            store.get("memo", ""),
            store.get("notes", ""),
            "Y" if store.get("contact_email") else "N",
            "Y" if store.get("contact_linkedin") else "N",
            "Y" if store.get("contact_remember") else "N",
            "Y" if store.get("contact_intro") else "N",
            "Y" if store.get("showOnMap", True) else "N",
            last_date,
            last_result,
            len(visits),
            "Y" if store.get("starred", False) else "N",
            store.get("created_at", ""),
        ])

    for col_idx, header in enumerate(headers1, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) * 2.5, 12)

    # 시트2: 전체 방문기록
    ws2 = wb.create_sheet("전체 방문기록")
    headers2 = ["가맹점명", "주소", "방문일", "방문결과", "방문메모"]
    ws2.append(headers2)
    style_header_row(ws2, len(headers2))

    for store in stores:
        for visit in store.get("visits", []):
            ws2.append([
                store.get("name", ""),
                store.get("address", ""),
                visit.get("date", ""),
                visit.get("result", ""),
                visit.get("memo", ""),
            ])

    for col_idx, header in enumerate(headers2, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) * 2.5, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stores_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ══════════════════════════════════════════
#  API 라우트 - JSON 백업 다운로드
# ══════════════════════════════════════════

@app.route("/api/backup")
@login_required
def backup_json():
    """Firestore 데이터를 JSON으로 다운로드"""
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my_stores = [s for s in stores if s.get("teamName") == team]

    output = json.dumps(my_stores, ensure_ascii=False, indent=2)
    filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    response = Response(
        output,
        mimetype="application/json; charset=utf-8",
    )
    response.headers["Content-Disposition"] = (
        f"attachment; filename=\"{filename}\""
    )
    return response


# ── 웹사이트 마이그레이션 (일회성) ──
@app.route("/api/migrate-websites", methods=["POST"])
@login_required
def migrate_websites():
    """기존 가맹점에 웹사이트 URL 일괄 추가"""
    user = get_current_user()
    if user.get("username") != "leesk0130":
        return jsonify({"error": "권한 없음"}), 403

    WEBSITE_MAP = {
        "공유어장": "https://padobox.kr",
        "파도상자": "https://padobox.kr",
        "해주세요": "https://haejuseyo.com",
        "라이펫": "https://lifet.co.kr",
        "아레스3": "https://honestflower.kr",
        "아니스트플라워": "https://honestflower.kr",
        "원모먼트": "https://www.1moment.co.kr",
        "플링캐스트": "https://www.plingcast.com",
        "쓰리랩스": "https://marketbang.kr",
        "마켓뱅": "https://marketbang.kr",
    }

    stores = load_stores()
    updated = 0
    for store in stores:
        if store.get("website"):
            continue
        name = store.get("name", "")
        for key, url in WEBSITE_MAP.items():
            if key in name:
                store["website"] = url
                save_store(store)
                updated += 1
                break

    return jsonify({"updated": updated})


# ── 캘린더 메모 API ──
_notes_cache = {"data": None, "ts": 0}

def load_notes():
    now = time.time()
    if _notes_cache["data"] is not None and (now - _notes_cache["ts"]) < 30:
        return _notes_cache["data"]
    result = fs_get_collection("calendar_notes")
    _notes_cache["data"] = result
    _notes_cache["ts"] = now
    return result

def _invalidate_notes_cache():
    _notes_cache["data"] = None
    _notes_cache["ts"] = 0


@app.route("/api/calendar-notes", methods=["GET"])
@login_required
def get_calendar_notes():
    user = get_current_user()
    team = user.get("teamName", "")
    notes = [n for n in load_notes() if n.get("teamName") == team]
    return jsonify(notes)


@app.route("/api/calendar-notes", methods=["POST"])
@login_required
def add_calendar_note():
    user = get_current_user()
    data = request.get_json()
    note = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "memo": data.get("memo", ""),
        "teamName": user.get("teamName", ""),
        "author": user.get("name", ""),
        "created_at": datetime.now().isoformat(),
    }
    fs_set_doc("calendar_notes", note["id"], note)
    _invalidate_notes_cache()
    return jsonify(note)


@app.route("/api/calendar-notes/<note_id>", methods=["DELETE"])
@login_required
def delete_calendar_note(note_id):
    user = get_current_user()
    team = user.get("teamName", "")
    notes = load_notes()
    for n in notes:
        if n["id"] == note_id and n.get("teamName") == team:
            fs_delete_doc("calendar_notes", note_id)
            _invalidate_notes_cache()
            return jsonify({"ok": True})
    return jsonify({"error": "메모를 찾을 수 없습니다."}), 404


# ══════════════════════════════════════════
#  API 라우트 - 통계 (주간/전환율/요일별)
# ══════════════════════════════════════════

STATUS_MAP = {
    '미방문': '미컨택',
    '부재중': '미컨택',
    '계약성사': '미팅완료',
    '실패': '미컨택',
    '기타': '미컨택',
}


@app.route("/api/stats/weekly", methods=["GET"])
@login_required
def stats_weekly():
    """이번 주(월~일) vs 지난 주 활동 건수 비교"""
    from datetime import timedelta
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]

    today = datetime.now().date()
    # 이번 주 월요일 (weekday(): 월=0)
    this_monday = today - timedelta(days=today.weekday())
    last_monday = this_monday - timedelta(days=7)

    def empty_bucket():
        return {"total": 0, "명함전달": 0, "미팅대기": 0, "미팅완료": 0}

    this_week = empty_bucket()
    last_week = empty_bucket()

    for s in my:
        for v in (s.get("visits") or []):
            try:
                vd = datetime.strptime(v.get("date", ""), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            result = v.get("result", "")
            if this_monday <= vd <= this_monday + timedelta(days=6):
                this_week["total"] += 1
                if result in this_week:
                    this_week[result] += 1
            elif last_monday <= vd <= last_monday + timedelta(days=6):
                last_week["total"] += 1
                if result in last_week:
                    last_week[result] += 1

    return jsonify({"thisWeek": this_week, "lastWeek": last_week})


@app.route("/api/stats/conversion", methods=["GET"])
@login_required
def stats_conversion():
    """전환율 퍼널 — 현재 각 상태별 가맹점 수"""
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]

    funnel = {"미컨택": 0, "명함전달": 0, "미팅대기": 0, "미팅완료": 0}

    for s in my:
        vs = s.get("visits") or []
        if not vs:
            funnel["미컨택"] += 1
            continue
        latest = max(vs, key=lambda v: v.get("date", ""))
        raw = latest.get("result", "") or ""
        status = STATUS_MAP.get(raw, raw)
        if status in funnel:
            funnel[status] += 1
        else:
            funnel["미컨택"] += 1

    return jsonify(funnel)


@app.route("/api/stats/activity-by-day", methods=["GET"])
@login_required
def stats_activity_by_day():
    """최근 7일 요일별 활동 건수"""
    from datetime import timedelta
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]

    today = datetime.now().date()
    day_names = ["월", "화", "수", "목", "금", "토", "일"]

    # 최근 7일 날짜 목록 (오래된 순)
    days = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        days.append({"date": d.strftime("%Y-%m-%d"), "day": day_names[d.weekday()], "count": 0})

    date_set = {d["date"] for d in days}

    for s in my:
        for v in (s.get("visits") or []):
            vdate = v.get("date", "")
            if vdate in date_set:
                for d in days:
                    if d["date"] == vdate:
                        d["count"] += 1
                        break

    return jsonify(days)


# ══════════════════════════════════════════
#  API 라우트 - 영업 히스토리 스냅샷
# ══════════════════════════════════════════

VALID_STATUSES = ['미컨택', '명함전달', '미팅대기', '미팅완료']

def _get_store_status(store):
    """매장의 현재 영업 상태 판별"""
    visits = store.get("visits") or []
    if not visits:
        return "미컨택"
    latest = max(visits, key=lambda v: v.get("date", ""))
    raw = latest.get("result", "") or ""
    mapped = STATUS_MAP.get(raw, raw)
    return mapped if mapped in VALID_STATUSES else "미컨택"


def _build_snapshot(stores, period_start, period_end):
    """주어진 기간의 스냅샷 데이터 생성"""
    total = len(stores)
    status_counts = {s: 0 for s in VALID_STATUSES}
    district_counts = {}
    contact_counts = {"email": 0, "linkedin": 0, "remember": 0, "intro": 0}
    activities_in_period = 0
    new_stores_in_period = 0
    activity_details = []  # 기간 내 활동 상세

    for s in stores:
        st = _get_store_status(s)
        status_counts[st] = status_counts.get(st, 0) + 1

        district = s.get("district", "기타")
        district_counts[district] = district_counts.get(district, 0) + 1

        if s.get("contact_email"):
            contact_counts["email"] += 1
        if s.get("contact_linkedin"):
            contact_counts["linkedin"] += 1
        if s.get("contact_remember"):
            contact_counts["remember"] += 1
        if s.get("contact_intro"):
            contact_counts["intro"] += 1

        # 등록일이 기간 내인지
        created = s.get("created_at", "")
        if created and period_start <= created[:10] <= period_end:
            new_stores_in_period += 1

        # 기간 내 활동 집계
        for v in (s.get("visits") or []):
            vdate = v.get("date", "")
            if vdate and period_start <= vdate[:10] <= period_end:
                activities_in_period += 1
                activity_details.append({
                    "store_name": s.get("name", ""),
                    "date": vdate[:10],
                    "result": STATUS_MAP.get(v.get("result", ""), v.get("result", "")),
                    "memo": v.get("memo", "")
                })

    contacted = total - status_counts.get("미컨택", 0)
    conversion_rate = round((contacted / total * 100), 1) if total > 0 else 0.0

    return {
        "total_stores": total,
        "status": status_counts,
        "districts": district_counts,
        "contacts": contact_counts,
        "activities_in_period": activities_in_period,
        "new_stores": new_stores_in_period,
        "contacted": contacted,
        "conversion_rate": conversion_rate,
        "activity_details": activity_details[:100],  # 최대 100건
    }


@app.route("/api/snapshots", methods=["GET"])
@login_required
def list_snapshots():
    """저장된 스냅샷 목록 조회"""
    user = get_current_user()
    team = user.get("teamName", "")
    snap_type = request.args.get("type", "")  # weekly / monthly / ''

    docs = fs_get_collection("snapshots")
    result = []
    for d in docs:
        if d.get("teamName") != team:
            continue
        if snap_type and d.get("type") != snap_type:
            continue
        # activity_details는 목록에서 제외 (무거움)
        item = {k: v for k, v in d.items() if k != "activity_details"}
        result.append(item)

    result.sort(key=lambda x: x.get("period_end", ""), reverse=True)
    return jsonify(result)


@app.route("/api/snapshots/<snap_id>", methods=["GET"])
@login_required
def get_snapshot(snap_id):
    """스냅샷 상세 조회 (activity_details 포함)"""
    user = get_current_user()
    team = user.get("teamName", "")

    docs = fs_get_collection("snapshots")
    for d in docs:
        if d.get("id") == snap_id and d.get("teamName") == team:
            return jsonify(d)
    return jsonify({"error": "스냅샷을 찾을 수 없습니다"}), 404


@app.route("/api/snapshots", methods=["POST"])
@login_required
def create_snapshot():
    """스냅샷 생성 (주별 또는 월별)"""
    from datetime import timedelta
    import calendar as cal_mod

    user = get_current_user()
    team = user.get("teamName", "")
    body = request.get_json(force=True)
    snap_type = body.get("type", "weekly")  # weekly / monthly
    target_date = body.get("date", "")  # YYYY-MM-DD (해당 주/월 기준일)

    if not target_date:
        target_date = datetime.now().strftime("%Y-%m-%d")

    try:
        dt = datetime.strptime(target_date, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "날짜 형식 오류 (YYYY-MM-DD)"}), 400

    if snap_type == "weekly":
        # 해당 주의 월~일
        monday = dt - timedelta(days=dt.weekday())
        sunday = monday + timedelta(days=6)
        period_start = monday.strftime("%Y-%m-%d")
        period_end = sunday.strftime("%Y-%m-%d")
        iso_cal = dt.isocalendar()
        period_label = f"{iso_cal[0]}-W{iso_cal[1]:02d}"
    elif snap_type == "monthly":
        period_start = dt.strftime("%Y-%m-01")
        last_day = cal_mod.monthrange(dt.year, dt.month)[1]
        period_end = dt.strftime(f"%Y-%m-{last_day:02d}")
        period_label = dt.strftime("%Y-%m")
    else:
        return jsonify({"error": "type은 weekly 또는 monthly"}), 400

    # 중복 체크
    existing = fs_get_collection("snapshots")
    for d in existing:
        if d.get("teamName") == team and d.get("period") == period_label and d.get("type") == snap_type:
            # 이미 존재하면 덮어쓰기
            snap_id = d.get("id")
            stores = load_stores()
            my = [s for s in stores if s.get("teamName") == team]
            snapshot_data = _build_snapshot(my, period_start, period_end)

            doc = {
                "id": snap_id,
                "teamName": team,
                "type": snap_type,
                "period": period_label,
                "period_start": period_start,
                "period_end": period_end,
                "created_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "created_by": user.get("username", ""),
            }
            doc.update(snapshot_data)
            fs_set_doc("snapshots", snap_id, doc)
            return jsonify({"message": "스냅샷 업데이트 완료", "id": snap_id, "snapshot": doc})

    # 새로 생성
    snap_id = f"snap_{uuid.uuid4().hex[:12]}"
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]
    snapshot_data = _build_snapshot(my, period_start, period_end)

    doc = {
        "id": snap_id,
        "teamName": team,
        "type": snap_type,
        "period": period_label,
        "period_start": period_start,
        "period_end": period_end,
        "created_at": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "created_by": user.get("username", ""),
    }
    doc.update(snapshot_data)
    fs_set_doc("snapshots", snap_id, doc)
    return jsonify({"message": "스냅샷 생성 완료", "id": snap_id, "snapshot": doc}), 201


@app.route("/api/snapshots/<snap_id>", methods=["DELETE"])
@login_required
def delete_snapshot(snap_id):
    """스냅샷 삭제"""
    user = get_current_user()
    team = user.get("teamName", "")

    docs = fs_get_collection("snapshots")
    for d in docs:
        if d.get("id") == snap_id and d.get("teamName") == team:
            fs_delete_doc("snapshots", snap_id)
            return jsonify({"message": "삭제 완료"})
    return jsonify({"error": "스냅샷을 찾을 수 없습니다"}), 404


@app.route("/api/snapshots/export", methods=["GET"])
@login_required
def export_snapshots():
    """스냅샷 히스토리를 엑셀로 내보내기"""
    user = get_current_user()
    team = user.get("teamName", "")
    snap_type = request.args.get("type", "")

    docs = fs_get_collection("snapshots")
    snapshots = [d for d in docs if d.get("teamName") == team]
    if snap_type:
        snapshots = [d for d in snapshots if d.get("type") == snap_type]
    snapshots.sort(key=lambda x: x.get("period_start", ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영업 히스토리"

    headers = ["기간", "유형", "시작일", "종료일", "전체 가맹점", "미컨택", "명함전달",
               "미팅대기", "미팅완료", "컨택율(%)", "기간내 활동", "신규 등록", "생성일"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, s in enumerate(snapshots, 2):
        status = s.get("status", {})
        if isinstance(status, str):
            try:
                status = json.loads(status)
            except:
                status = {}
        ws.cell(row=row_idx, column=1, value=s.get("period", ""))
        ws.cell(row=row_idx, column=2, value="주별" if s.get("type") == "weekly" else "월별")
        ws.cell(row=row_idx, column=3, value=s.get("period_start", ""))
        ws.cell(row=row_idx, column=4, value=s.get("period_end", ""))
        ws.cell(row=row_idx, column=5, value=s.get("total_stores", 0))
        ws.cell(row=row_idx, column=6, value=status.get("미컨택", 0))
        ws.cell(row=row_idx, column=7, value=status.get("명함전달", 0))
        ws.cell(row=row_idx, column=8, value=status.get("미팅대기", 0))
        ws.cell(row=row_idx, column=9, value=status.get("미팅완료", 0))
        ws.cell(row=row_idx, column=10, value=s.get("conversion_rate", 0))
        ws.cell(row=row_idx, column=11, value=s.get("activities_in_period", 0))
        ws.cell(row=row_idx, column=12, value=s.get("new_stores", 0))
        ws.cell(row=row_idx, column=13, value=s.get("created_at", ""))

    # 열 너비 조절
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"영업히스토리_{team}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ══════════════════════════════════════════
#  API 라우트 - 팀원별 활동 현황
# ══════════════════════════════════════════

@app.route("/api/stats/by-member", methods=["GET"])
@login_required
def stats_by_member():
    """팀원별 활동 건수, 최근 활동일 집계"""
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my = [s for s in stores if s.get("teamName") == team]

    # 팀원 목록
    users = load_users()
    team_members = [u for u in users if u.get("teamName") == team and u.get("isApproved")]

    member_stats = {}
    for u in team_members:
        member_stats[u["id"]] = {
            "id": u["id"],
            "name": u.get("name", "알 수 없음"),
            "total_activities": 0,
            "last_activity_date": "",
            "this_month_activities": 0,
            "status_counts": {"미컨택": 0, "명함전달": 0, "미팅대기": 0, "미팅완료": 0},
        }

    today = datetime.now().date()
    this_month_start = today.replace(day=1).strftime("%Y-%m-%d")

    for s in my:
        for v in (s.get("visits") or []):
            v_author = v.get("author", "") or v.get("created_by", "")
            v_date = v.get("date", "")

            matched_id = None
            if v_author:
                for uid, ms in member_stats.items():
                    if ms["name"] == v_author:
                        matched_id = uid
                        break

            if not matched_id and len(team_members) == 1:
                matched_id = team_members[0]["id"]

            if not matched_id:
                if "__unassigned__" not in member_stats:
                    member_stats["__unassigned__"] = {
                        "id": "__unassigned__",
                        "name": "미지정",
                        "total_activities": 0,
                        "last_activity_date": "",
                        "this_month_activities": 0,
                        "status_counts": {"미컨택": 0, "명함전달": 0, "미팅대기": 0, "미팅완료": 0},
                    }
                matched_id = "__unassigned__"

            ms = member_stats[matched_id]
            ms["total_activities"] += 1

            if v_date and v_date > ms["last_activity_date"]:
                ms["last_activity_date"] = v_date

            if v_date and v_date >= this_month_start:
                ms["this_month_activities"] += 1

            result = v.get("result", "미컨택") or "미컨택"
            mapped = STATUS_MAP.get(result, result)
            if mapped in ms["status_counts"]:
                ms["status_counts"][mapped] += 1

    result_list = sorted(member_stats.values(), key=lambda x: x["total_activities"], reverse=True)
    return jsonify(result_list)


# ══════════════════════════════════════════
#  API 라우트 - 활동 로그 조회
# ══════════════════════════════════════════

@app.route("/api/activity-logs", methods=["GET"])
@login_required
def get_activity_logs():
    """팀 활동 로그 조회 (최신순). ?limit=50&offset=0"""
    user = get_current_user()
    team = user.get("teamName", "")
    limit = int(request.args.get("limit", 50))
    offset = int(request.args.get("offset", 0))

    all_logs = fs_get_collection("activity_logs")
    team_logs = [l for l in all_logs if l.get("teamName") == team]
    team_logs.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    paginated = team_logs[offset:offset + limit]
    return jsonify({"logs": paginated, "total": len(team_logs)})


@app.route("/api/inactive-stores", methods=["GET"])
@login_required
def get_inactive_stores():
    """14일 이상 활동 없는 가맹점 목록"""
    user = get_current_user()
    team = user.get("teamName", "")
    stores = load_stores()
    my_stores = [s for s in stores if s.get("teamName") == team]

    from datetime import timedelta
    cutoff = datetime.now() - timedelta(days=14)
    inactive = []
    for s in my_stores:
        visits = s.get("visits") or []
        if visits:
            latest = max(visits, key=lambda v: v.get("date", ""))
            last_date_str = latest.get("date", "")
        else:
            last_date_str = s.get("created_at", "")

        if not last_date_str:
            inactive.append({"id": s["id"], "name": s.get("name", ""), "last_activity": ""})
            continue

        try:
            last_dt = datetime.fromisoformat(last_date_str.replace("Z", "+00:00").split("+")[0].split("T")[0])
        except Exception:
            try:
                last_dt = datetime.strptime(last_date_str[:10], "%Y-%m-%d")
            except Exception:
                inactive.append({"id": s["id"], "name": s.get("name", ""), "last_activity": last_date_str})
                continue

        if last_dt < cutoff:
            days_ago = (datetime.now() - last_dt).days
            inactive.append({"id": s["id"], "name": s.get("name", ""), "last_activity": last_date_str, "days_ago": days_ago})

    inactive.sort(key=lambda x: x.get("days_ago", 9999), reverse=True)
    return jsonify({"stores": inactive, "count": len(inactive)})


# ══════════════════════════════════════════
#  카카오 Static Map 프록시
# ══════════════════════════════════════════

@app.route("/api/kakao-staticmap")
@login_required
def kakao_staticmap():
    """카카오 Static Map 프록시 — API 키 숨김"""
    lat = request.args.get("lat", "")
    lng = request.args.get("lng", "")
    width = request.args.get("width", "300")
    height = request.args.get("height", "200")
    level = request.args.get("level", "3")

    url = f"https://dapi.kakao.com/v2/maps/staticmap?center={lng},{lat}&level={level}&w={width}&h={height}&maptype=roadmap&marker=color:red|{lng},{lat}"
    headers = {"Authorization": f"KakaoAK {KAKAO_API_KEY}"}
    resp = requests.get(url, headers=headers, timeout=10)
    return Response(resp.content, content_type=resp.headers.get('content-type', 'image/png'))


# ══════════════════════════════════════════
#  법적 페이지 라우트
# ══════════════════════════════════════════

@app.route("/privacy")
def privacy_page():
    return render_template("privacy.html")

@app.route("/terms")
def terms_page():
    return render_template("terms.html")

@app.route("/licenses")
def licenses_page():
    return render_template("licenses.html")


# ══════════════════════════════════════════
#  Keep-alive (Render Free tier 슬립 방지)
# ══════════════════════════════════════════

@app.route("/health")
def health_check():
    return "ok", 200

def _keep_alive():
    """10분마다 자기 서버에 핑을 보내 Render 슬립 방지"""
    import urllib.request
    url = "https://point3-salesmap.onrender.com/health"
    while True:
        time.sleep(600)  # 10분
        try:
            urllib.request.urlopen(url, timeout=10)
        except Exception:
            pass

# ── 초기화 + 서버 실행 ──
ensure_superadmin()

# keep-alive 스레드 시작 (gunicorn에서도 동작)
_keep_alive_thread = threading.Thread(target=_keep_alive, daemon=True)
_keep_alive_thread.start()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=os.environ.get("FLASK_DEBUG", "true").lower() == "true")
